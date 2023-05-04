import os
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
BASE_DIR = os.path.dirname(os.path.abspath(__file__)).strip('\\src')

######################### Sorting Log File ############################
def sort_logs(log_file):
    f = open(log_file)
    lines = f.readlines()
    f.close()

    with open(log_file, 'w') as file:
        lines = sorted(lines, key=lambda string: int(string.split(',')[0]))
        file.writelines(lines)

def highlight_logs(temp_file, file_name, log_file):
    sort_logs(log_file)

    workbook = openpyxl.load_workbook(temp_file)
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[1]]
    max_cols = worksheet.max_column

    with open(log_file) as file:
        lines = file.readlines()
        for line in lines:
            row = int(line.split(',')[0])
            for cell in worksheet[row]:
                cell.fill = fill
    workbook.save(file_name)