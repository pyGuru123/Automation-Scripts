import os
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, GradientFill, Alignment, Font
from openpyxl.utils import get_column_letter
from typing import BinaryIO

fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
fill2 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill3 = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
fill4 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BASE_DIR = os.path.dirname(os.path.abspath(__file__)).strip("\\src")


######################### Sorting Log File ############################
def sort_logs(log_file: str) -> None:
    f = open(log_file)
    lines = f.readlines()
    f.close()

    with open(log_file, "w") as file:
        lines = sorted(lines, key=lambda string: int(string.split("@")[0]))
        file.writelines(lines)


def highlight_logs(excel_file: BinaryIO, filename: str, log_file: str) -> None:
    sort_logs(log_file)

    workbook = openpyxl.load_workbook(excel_file)
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[1]]
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    worksheet.insert_cols(max_col + 1)

    column_letter = get_column_letter(max_col + 1)
    worksheet.column_dimensions[column_letter].width = 100

    cell = worksheet.cell(row=1, column=max_col + 1)
    cell.value = "Failure Log"
    cell.font = Font(color="00FFFFFF", size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = GradientFill(stop=("FF0000", "DF1B1B"))

    with open(log_file) as file:
        lines = file.readlines()
        for line in lines:
            if line.strip() != "":
                row_num, text = line.split(":")
                row, col = list(map(int, row_num.split("@")))

                cell = worksheet.cell(row=int(row), column=col)
                cell.fill = fill

                col_value = worksheet.cell(row=int(row), column=max_col + 1)
                if col_value.value == None:
                    col_value.value = ""
                worksheet.cell(row=int(row), column=max_col + 1).value = (
                    col_value.value + text + ", "
                )

                if "exclusion" in col_value.value:
                    cell.fill = fill2
                elif "people per company" in col_value.value:
                    cell.fill = fill3
                elif row > max_row-1 and col < max_col:
                    cell.fill = fill4

                col_value.fill = fill

    workbook.save(filename)
    workbook.close()
