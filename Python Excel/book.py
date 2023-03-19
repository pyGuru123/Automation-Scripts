from openpyxl.workbook import Workbook 
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

# creating new worksheet

ws1 = wb.create_sheet('NewSheet', 1)
ws2 = wb.create_sheet('sheet 2', 0)

ws.title = 'MySheet'

# print(wb.sheetnames)

wb2 = load_workbook("files/regions.xlsx")
# new_sheet = wb2.create_sheet("NewSheet")
print(wb2.sheetnames)
sheet = wb2.active 
print(sheet)

cell = sheet['A1']
print(cell, cell.value)
cell.value = "Batman"
wb2.save("files/modified_regions.xlsx")