from openpyxl import load_workbook

wb = load_workbook("files/regions.xlsx")
ws = wb.active

cell_range = ws['A1':'C1']
print([cell for cell in cell_range])

column = ws['C']
col_range = ws['A':'C']

row = ws[1]
row_range = ws[1:5]
print(row_range)

for row in ws.iter_rows(min_row=1, max_row=2, max_col=3, values_only=True):
    for cell in row:
        print(cell)