from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

# cell merge
ws.append(["Table (1-20)"])
ws.merge_cells("A1:J1")
# ws.unmerge_cells("A1:J1")

# gtradient fill and allignment
cell = ws['A1']
cell.font = Font(color="0000FF00", size=15, bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = GradientFill(stop=('000000', 'FFFFFF'))

# filling cells with values
for row in range(1, 21):
	ws.append([row*i for i in range(1, 11)])

# swetting value and color
cell = ws["B2"]
cell.font = Font(color='00FF0000', size=12, italic=True)
cell.alignment = Alignment(horizontal="right", vertical="bottom")
cell.value = "Something"
cell.fill = GradientFill(stop=('FF0000', '0000FF'))

# border color and solid colors
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True)
bd = Side(style="thick", color="00FF00")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill('solid', fgColor='FFFF00')

count = 0
for col in ws.iter_cols(min_col=1, min_row=2, max_col=10, max_row=30):
	col[count].style = highlight
	count += 1

wb.save("files/tables.xlsx")